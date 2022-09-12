class Courses_attachment:
    cell_names = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']
    from openpyxl import load_workbook
    from datetime import datetime
    now = datetime.now()
    current_date = now.strftime("%Y%m%d")
    def unban(self,sheet):
        i,counter = 2,0
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        x = str(input('Enter ID of that student to ban: '))
        while True:
            if ws['A'+str(i)].value != None: 
                counter += 1
            if ws['A'+str(i)].value == None:
                break
            i+=1
        for i in range(counter):
            if (ws['B'+str(i+2)].value == x) and (ws['C'+str(i+2)].value == True):
                ws['C'+str(i+2)].value = False
                print('The student with ID',x,'has been unbanned successfully ')
                return 
            if (ws['B'+str(i+2)].value == x) and (ws['C'+str(i+2)].value == False):
                print('This student is already unbanned. \n')
                return
        print('No student with ID is enrolled.\n')
    def ban(self,sheet):
        i,counter = 2,0
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        x = str(input('Enter ID of that student to ban: '))
        while True:
            if ws['A'+str(i)].value != None: 
                counter += 1
            if ws['A'+str(i)].value == None:
                break
            i+=1
        for i in range(counter):
            if (ws['B'+str(i+2)].value == x) and (ws['C'+str(i+2)].value == False):
                ws['C'+str(i+2)].value = True
                print('The student with ID',x,'is banned successfully!\n ')
                return 
            if (ws['B'+str(i+2)].value == x) and (ws['C'+str(i+2)].value == True):
                print('This student is already banned.\n')
                return
        print('No student with ID is enrolled in this course.\n')
    def enrol_a_student(self,sheet):
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        row_count = (ws.max_row)+1
        ws['A'+str(row_count)] = input('enter name to enrol the student : ')
        ws['B'+str(row_count)] = input('enter ID of student:  ')
        ws['C'+str(row_count)] = False
        wb.save('AI.xlsx')
    def attendence(self,sheet):
        counter = 0
        i = 2
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        column_count = (ws.max_column)
        cell_name = Courses_attachment.cell_names[column_count]
        while True:
            if ws['A'+str(i)].value != None: 
                counter += 1
            if ws['A'+str(i)].value == None:
                break
            i+=1
        for i in range(counter):
            if ws['C'+str(i+2)].value == False:
                print('Is',ws['A'+str(i+2)].value,'present or not')
                x = input('enter P or A to mark attendence: ')
                ws[cell_name+str(i+2)] = x
        ws[cell_name+str(1)] =  Courses_attachment.current_date
        wb.save('AI.xlsx')
        print('Attendence has been marked successfully marked.\n')
        Courses_attachment.generate_summary(self,self.file_name,counter,cell_name,sheet)
    def print_status_of_student_for_today(self,sheet):
        counter = 0
        i = 2
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        column_count = (ws.max_column)
        cell = (Courses_attachment.cell_names[column_count-1])
        while True:
            if ws['A'+str(i)].value != None: 
                counter += 1
            if ws['A'+str(i)].value == None:
                break
            i+=1
        for i in range(counter-1):
            print(ws['A'+str(i+2)].value,'was: ',ws[cell+str(i+2)].value,'\n')
    def generate_summary(self, a,counter,cell,sheet):
        present, absent = 0,0
        wb = Courses_attachment.load_workbook(a)
        ws = wb[sheet]
        for i in range(counter):
            if ws[cell+str(i+2)].value == 'P':
                present += 1
            if ws[cell+str(i+2)].value == 'A':
                absent += 1
        ws[cell+str(41)] = present
        ws[cell+str(42)] = absent
        ws[cell+str(43)] = (present/counter)*100
        ws[cell+str(44)] = (absent/counter)*100
        wb.save('AI.xlsx')
    def check_attendence_call(self,sheet):
        self.attendence = False
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        column_count = (ws.max_column)
        cell = (Courses_attachment.cell_names[column_count-1])+'1'
        if int(ws[cell].value) != int(Courses_attachment.current_date):
            self.attendence = True
        if self.attendence == True:
            Courses_attachment.attendence(self,sheet)
        else:
            print('dont worry Sir your todays attendence has been marked!\n')
            print('If you want to check the status of todays attendence enter 1:\nelse enter 0:\n')
            y = int(input('Choose 1 or 0: '))
            if y == 1:
                Courses_attachment.print_status_of_student_for_today(self,sheet)
    def give_summary_of_a_students(self,sheet):
        self.presence = False
        present, absent,a = 0,0,0
        self.file_name = 'AI.xlsx'
        wb = Courses_attachment.load_workbook(self.file_name)
        ws = wb[sheet]
        column_count = (ws.max_column)
        x = str(input('Enter ID of that student to show summary: '))
        for i in range(0,40):
            if ws['B'+str(i+2)].value == x:
                self.presence = True
                a = i+2
        if self.presence == False:
            print('No student with this ID\n')
            return 
        for i in range(column_count-3):
            if ws[(Courses_attachment.cell_names[i+3])+str(a)].value == 'P':
                present += 1
            if ws[(Courses_attachment.cell_names[i+3])+str(a)].value == 'A':
                absent += 1
        print('Total sessions held are',column_count-3,'\nTotal sessions taken by',x,'are',present,'\nAbsents of',x,'are',absent,'\n')
import texttable
tableObj = texttable.Texttable()
tableObj.add_rows([
		["Courses", "Instructor", "Student ID"],
		["Artificial Intelligence", "Dr Junaid Akhter", "NIM-AI-"],
		["Cyber Secuirty", "Dr Saad Ali Malik", "NIM-CS-"],
		["Machine Learning", "Dr Malik Jahan", "NIM-ML-"],
		])

print(tableObj.draw())
print('Sir, summary of each will create at ROW:41 to 44, this I have done due to space left for entolnment of students.\n')                     
sheets = ['AI','Cyber Secuirty','Machine Learning']  
DJ = Courses_attachment()
print(['AI','Cyber Secuirty','Machine Learning'],'are the courses.\nattendence of which course do you ant to call?\n')
while True:
    q = str(input('Choose your course to mark attendence: '))
    sheet = sheets[['AI','Cyber Secuirty','Machine Learning'].index(q)]
    a = int(input('\nfor attendence enter 0: \nfor enrolment enter 1: \nto see summary of any student enter 2:\nto ban any student enter 3:''\nto unban any student enter 4:\nto logout enter 5:'))
    if a == 0:
        DJ.check_attendence_call(sheet)
    elif  a== 1:
        DJ.enrol_a_student(sheet)
    elif a == 2:
        DJ.give_summary_of_a_students(sheet)
    elif a == 3:
         DJ.ban(sheet)
    elif a == 4:
        DJ.unban(sheet)
    elif a == 5:
        break
