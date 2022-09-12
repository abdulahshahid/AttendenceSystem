class Namal_University:
    from datetime import datetime
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    current_date = now.strftime("%d/%m/%Y")
    from openpyxl import load_workbook
    import pandas as pd
    '''This function is the property of administration to signup new students.'''
    def sign_up():
        wb1 = Namal_University.load_workbook('Students_attendene.xlsx')
        ws1 = wb1['Sheet1']
        rows_count = (ws1.max_row) + 1
        wb = Namal_University.load_workbook('Students_Record_signup.xlsx')
        ws = wb['Sheet1']
        row_count = (ws.max_row)+1
        ws['A'+str(row_count)] = input('enter first name: ')
        ws['B'+str(row_count)] = input('enter last name: ')
        ws['C'+str(row_count)] = input('enter city: ')
        ID = 'NIM-BS-'+str(row_count)  
        ws['D'+str(row_count)] = ID 
        ws1['A'+str(rows_count)] = ID
        password = input('Enter password: ')
        ws['E'+str(row_count)] = password
        ws1['B'+str(rows_count)] = password
        ws['F'+str(row_count)] = Namal_University.current_time
        ws['G'+str(row_count)] = Namal_University.current_date
        wb.save('Students_Record_signup.xlsx')
        wb1.save('Students_attendene.xlsx')
    '''This function is the property of administrationto show data of student'''
    def show_data():
        a = 0
        df = Namal_University.pd.read_excel('Students_Record_signup.xlsx')
        data = df.values
        x = input('Data of which you want to know? Enter the Enerlnment ID of that student. ')
        for i in range(len(df)):
            if data[i,3] == str(x):
                print('name of that sudent is',str(data[i,0])+" "+str(data[i,1]), '\nThis student belongs to: ',data[i,2],'\nThis student was enrolled at: ',data[i,6])
                a +=1
        if a == 0:
            print('No student with this id\n')

    '''This function is the property of administration to reset password of student'''
    def change_password():
        A = 0
        def change_password2(i,a):
            wb = Namal_University.load_workbook('Students_attendene.xlsx')
            ws= wb['Sheet1']
            ws['B'+str(i+2)] = a
            wb.save('Students_attendene.xlsx')
            print('Password changed successfully''\n')
        df = Namal_University.pd.read_excel('Students_Record_signup.xlsx')
        data = df.values
        wb = Namal_University.load_workbook('Students_Record_signup.xlsx')
        ws = wb['Sheet1']
        q = input('Enter ID of student to reset password: ')
        for i in range(len(df)):
            if data[i,3] == str(q):
                a = input('choose new password: ')
                ws['E'+str(i+2)] = a
                wb.save('Students_Record_signup.xlsx')
                change_password2(i,a)
                A+= 1
        if A == 0: print('NO student with ID found \n')
    '''this function is the property of administration to ban student'''

    def ban():
        A = 0
        df = Namal_University.pd.read_excel('Students_Record_signup.xlsx')
        data = df.values
        wb = Namal_University.load_workbook('Students_Record_signup.xlsx')
        ws = wb['Sheet1']
        x = input('enter ID of students to Ban/Unban: ')
        for i in range(len(df)):
            if (data[i,3] == str(x)) and (data[i,7] == False):
                ws['H'+str(i+2)] = True
                print('Student with ID',x,'is successfully banned.\n')
                A+=1
            if (data[i,3] == str(x)) and (data[i,7] == True):
                print('this student is already banned by adminintrative.\n ')
                A += 1
        if A==0:
            print('no student with ID FOUND \n')
        wb.save('Students_Record_signup.xlsx')
        
        '''this function is the property of administration to unban banned student'''
    def unban():
        A = 0
        df = Namal_University.pd.read_excel('Students_Record_signup.xlsx')
        data = df.values
        wb = Namal_University.load_workbook('Students_Record_signup.xlsx')
        ws = wb['Sheet1']
        x = input('enter ID of students to Ban/Unban: ')
        for i in range(len(df)):
            if (data[i,3] == str(x)) and (data[i,7] == True):
                ws['H'+str(i+2)] = False
                print('Student with ID',x,'is successfully unbanned.\n')
                A+=1
            if (data[i,3] == str(x)) and (data[i,7] == False):
                print('this student is already unbanned by adminintrative \n')
                A += 1
        if A==0:
            print('no student with ID FOUND ')
        wb.save('Students_Record_signup.xlsx')
        
    '''This function is the property of adminidtration to see entry and exit of late students'''
    def tell_late_entryandexit():
        wb = Namal_University.load_workbook('Students_attendene.xlsx')
        ws = wb['Sheet1']
        rows = ws.max_row
        for i in range(rows):
            if (str(ws['C'+str(i+1)].value))>'09:00:00' and (ws['C'+str(i+1)].value!= None):
                print((ws['A'+str(i+1)].value),'enters late at: ',ws['C'+str(i+1)].value)
            if (str(ws['D'+str(i+1)].value)>'17:00:00') and (ws['D'+str(i+1)].value!= None):
                print(str(ws['A'+str(i+1)].value), 'leaves late at: ',ws['D'+str(i+1)].value)
            if ws['D'+str(i+1)].value == None:
                print(ws['A'+str(i+1)].value, "enters but didn't leave")
            
    '''This function is public for entry of students.'''           
    def entry():
        a = 0
        wb = Namal_University.load_workbook('Students_attendene.xlsx')
        ws = wb['Sheet1']
        wb1 = Namal_University.load_workbook('Students_Record_signup.xlsx')
        ws1 = wb1['Sheet1']
        rows_count = ws.max_row
        x = input('Enter user ID to mark entry: ')
        for i in range(rows_count):
            if (ws['A'+str(i+1)].value == str(x)) and (ws1['H'+str(i+1)].value == False):
                ws['C'+str(i+2)] = Namal_University.current_time
                a+= 1
                print('Entry timing successfully marked\n')
            if (ws1['H'+str(i+1)].value == True):
                print('Banned by administration.')
                return 
        if (a == 0):
            print('No registration with this ID. Sign up first to mark entry!! ')
        wb.save('Students_attendene.xlsx')
        
    '''This function is public for exit of students.'''
    def exit():
        a = 0
        wb = Namal_University.load_workbook('Students_attendene.xlsx')
        ws = wb['Sheet1']
        df = Namal_University.pd.read_excel('Students_attendene.xlsx')
        data = df.values
        df1 = Namal_University.pd.read_excel('Students_Record_signup.xlsx')
        data1 = df1.values
        x = input('Enter user ID to exit: ')
        for i in range(len(df)):
            if ((data[i,0] == str(x)) and (data[i+1,0] != 'nan')) and (data1[i,7] == True):
                a+=1
                print('Mark your entry to exit !')
                Namal_University.entry()
            elif ((data[i,0] == str(x)) and (data[i+1,0] == 'nan')) and (data1[i,7] == True):
                a+=1
                ws['D'+str(i+2)] = Namal_University.current_time
                print('Your exit timing has been marked. \n')
            if (data1[i,7] == False):
                print('Banned by administration.\n')
        if a == 0:
            print('No registration with this ID. Sign up first to mark entry!! \n')
        wb.save('Students_attendene.xlsx')
    def enter_time():
        wb = Namal_University.load_workbook('Students_Record_signup.xlsx')
        ws = wb['Sheet1']
        rows = ws.max_row
        for i in range(rows):
            ws['F'+str(i+1)] = Namal_University.current_time
        wb.save('Students_Record_signup.xlsx')
Namal_University.enter_time()
while True:
    print('For Students options Enter 0\nFor admininstrative options enter 1\n')
    x = int(input('Make choice: '))
    if x== 0:
        y = int(input('do you want to mark entry or exit?\nFor entry enter 0, for exit enter 1  '))
        if y == 0:
            Namal_University.entry()
        if y == 1:
            Namal_University.exit()
    if x == 1:
        print('for using administrative options you will have to verify yourself. So enter username and password. ')
        print('\n')
        q = input('enter admin usernname: ')
        w = input('enter admin password: ')
        if (q == 'admin') and (w == 'admin285'):
            print('For Signingup enter 1\nTo reset password of any enter 2\nTo see the information of studenr enter 3\nTo ban an student enter 4\nTo unban an student enter 5\nTo see late entry and exit of students enter 6\nTo shutdown system enter 7 ')
            e = int(input('Make your choice: '))
            if e == 1:
                Namal_University.sign_up()
            elif e == 2:
                Namal_University.change_password()
            elif e == 3:
                Namal_University.show_data()
            elif e == 4:
                Namal_University.ban()
            elif e == 5:
                Namal_University.unban()
            elif e == 6:
                Namal_University.tell_late_entryandexit()
            elif e == 7:
                break
